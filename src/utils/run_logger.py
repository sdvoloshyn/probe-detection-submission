from __future__ import annotations

import csv
import json
import os
import shutil
import tempfile
from dataclasses import asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import torch
from filelock import FileLock, Timeout


RESULTS_DIR = Path.cwd() / "results"


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _atomic_append(csv_path: Path, rows: List[List[str]]) -> None:
    """Atomically append lines to a CSV file.
    Writes to a temp file and then renames into place to avoid partial writes.
    """
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    # Create tmp file in same directory to ensure atomic rename on POSIX
    fd, tmp_path = tempfile.mkstemp(prefix=csv_path.name + ".tmp_", dir=str(csv_path.parent))
    os.close(fd)
    try:
        # If dest exists, copy it to tmp first, then append
        if csv_path.exists():
            shutil.copy2(csv_path, tmp_path)
        # Append rows
        with open(tmp_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=",", lineterminator="\n")
            for r in rows:
                writer.writerow(r)
        # Atomic replace
        os.replace(tmp_path, csv_path)
    finally:
        # Clean up tmp if something failed before replace
        if os.path.exists(tmp_path) and not os.path.samefile(tmp_path, csv_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def append_with_headers(csv_path: Path, headers: List[List[str]], row: List[str]) -> None:
    """Append a single CSV row; write headers first if file missing.
    headers may contain one or two header rows.
    """
    lines: List[List[str]] = []
    if not csv_path.exists():
        for h in headers:
            lines.append(h)
    lines.append(row)
    _atomic_append(csv_path, lines)


def _fmt(val: Any, ndigits: int = 4) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        return f"{val:.{ndigits}f}"
    if isinstance(val, (int,)):
        return str(val)
    return str(val)


def format_float(x: Optional[float], ndigits: int = 4) -> str:
    return _fmt(x, ndigits=ndigits)


def format_time_min(seconds: Optional[float]) -> str:
    if seconds is None:
        return ""
    minutes = float(seconds) / 60.0
    return f"{minutes:.2f}"


def _iso_utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _git_commit_short() -> str:
    try:
        import subprocess
        root = Path.cwd()
        if (root / ".git").exists():
            out = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], cwd=str(root))
            return out.decode("utf-8").strip()
    except Exception:
        pass
    return ""


def _device_info() -> tuple[str, str]:
    if torch.cuda.is_available():
        try:
            name = torch.cuda.get_device_name(0)
        except Exception:
            name = "cuda"
        ver = torch.version.cuda or ""
        return name, ver
    return "cpu", ""


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa: F401
    except Exception as e:
        raise RuntimeError("openpyxl is required for XLSX export. Install with: pip install openpyxl") from e


def _append_xlsx_with_headers(
    xlsx_path: Path,
    headers: List[List[str]],
    row: List[str],
    merges: List[tuple[int, int, int, int]] | None = None,
) -> None:
    _ensure_openpyxl()
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if not xlsx_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        # write headers
        for r_idx, hdr in enumerate(headers, start=1):
            for c_idx, val in enumerate(hdr, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        # merges (row/col are 1-based)
        if merges:
            for r1, c1, r2, c2 in merges:
                ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        # style headers
        for r in range(1, len(headers) + 1):
            for c in range(1, len(headers[0]) + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(bold=True)
        # freeze header
        ws.freeze_panes = ws["A3"] if len(headers) >= 2 else ws["A2"]
        # column widths heuristic
        for idx in range(1, len(headers[0]) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 14
        # append first row
        ws.append(row)
        wb.save(xlsx_path)
    else:
        wb = load_workbook(xlsx_path)
        ws = wb.active
        ws.append(row)
        wb.save(xlsx_path)


class RunLogger:
    @staticmethod
    def append_compact(
        run_info: Dict[str, Any],
        folds_metrics: List[Dict[str, Any]] | None,
        aggregate: Dict[str, Any] | None,
        time_seconds: Optional[float],
        fold_index: Optional[int] = None,
    ) -> None:
        """Append a row to results/results_compact.csv.
        - Two header rows as specified
        - Support single split by placing value in fold1 and mean, std empty
        """
        ensure_dir(RESULTS_DIR)
        csv_path = RESULTS_DIR / "results_compact.csv"

        # COMPACT headers per new specification (reduced config cols; first col run_name)
        header_config_cols = [
            "run_name","timestamp","model","imgsz","batch_size","epochs_planned","n_augmentations","aug_profile","geom_profile"
        ]
        # Groups span: fold1..fold5 + mean (no std)
        header_group_labels = [
            ("F1@best", 6),
            ("mean_IoU@TP", 6),
            ("mAP50-95", 6),
            ("FPR@best", 6),
            ("best_conf", 6),
            ("best_epoch", 6),
        ]
        header1 = list(header_config_cols)
        for label, span in header_group_labels:
            header1 += [label] + [""] * (span - 1)
        header2 = list(header_config_cols)
        for _label, _span in header_group_labels:
            header2 += ["fold1","fold2","fold3","fold4","fold5","mean"]

        # Prepare folds arrays for metrics
        def collect(metric_key: str) -> tuple[List[Optional[float]], Optional[float], Optional[float]]:
            vals: List[Optional[float]] = [None, None, None, None, None]
            ms: List[float] = []
            if folds_metrics:
                for i, m in enumerate(folds_metrics[:5]):
                    v = m.get(metric_key)
                    if isinstance(v, (int, float)):
                        vals[i] = float(v)
                        ms.append(float(v))
            # aggregate mean/std if provided
            mean = None
            std = None
            if aggregate:
                mean = aggregate.get(f"{metric_key}_mean")
                std = aggregate.get(f"{metric_key}_std")
            # single-split: if no aggregate but a single value exists
            if mean is None and ms:
                if len(ms) == 1:
                    mean = ms[0]
                    std = None
            return vals, mean, std

        f1_vals, f1_mean, f1_std = collect("f1_best")
        iou_vals, iou_mean, iou_std = collect("mean_iou_tp_best")
        map_vals, map_mean, map_std = collect("map50_95")
        fpr_vals, fpr_mean, fpr_std = collect("fp_rate_best")
        conf_vals, conf_mean, conf_std = collect("best_conf")
        epoch_vals, epoch_mean, epoch_std = collect("epoch_best")

        def current_value(metric_key: str) -> Optional[float]:
            if not folds_metrics:
                return None
            try:
                v = folds_metrics[0].get(metric_key)
                if v is None:
                    return None
                return float(v)
            except Exception:
                return None

        cur_f1 = current_value("f1_best")
        cur_iou = current_value("mean_iou_tp_best")
        cur_map = current_value("map50_95")
        cur_fpr = current_value("fp_rate_best")
        cur_conf = current_value("best_conf")
        cur_epoch = current_value("epoch_best")

        # Config/identity row prefix (leave qualitative empty)
        run_name = str(run_info.get("run_name", run_info.get("run_id", "")))
        timestamp = run_info.get("timestamp") or _iso_utc_now()
        model = str(run_info.get("model", ""))
        imgsz = run_info.get("imgsz")
        batch_size = run_info.get("batch_size")
        epochs_planned = run_info.get("epochs_planned")
        n_augmentations = run_info.get("n_augmentations")

        row: List[str] = [
            run_name,
            timestamp,
            str(model),
            _fmt(imgsz, 0),
            _fmt(batch_size, 0),
            _fmt(epochs_planned, 0),
            _fmt(n_augmentations, 0),
            "",  # aug_profile
            "",  # geom_profile
        ]

        def flat(vals: List[Optional[float]], mean: Optional[float], std: Optional[float], digits: int = 4) -> List[str]:
            out = [format_float(v, ndigits=digits) if v is not None else "" for v in vals]
            out += [format_float(mean, ndigits=4), format_float(std, ndigits=4)]
            return out

        # Do not add metric blocks into initial row; we upsert per-fold and recompute mean in-sheet
        # epoch_best formatted as integers
        def _to_int(x: Optional[float]) -> Optional[float]:
            if x is None:
                return None
            try:
                return float(int(round(float(x))))
            except Exception:
                return None
        epoch_vals_int = [_to_int(v) for v in epoch_vals]
        epoch_mean_int = _to_int(epoch_mean)
        epoch_std_int = _to_int(epoch_std) if epoch_std is not None else None
        row += flat(epoch_vals_int, epoch_mean_int, epoch_std_int, 0)

        # Write/update XLSX with merged cells and single row per (run_name, timestamp)
        # Build merges once on file creation
        merges: List[tuple[int, int, int, int]] = []
        col = 1
        for idx in range(len(header_config_cols)):
            merges.append((1, col, 2, col))
            col += 1
        for _label, span in header_group_labels:
            merges.append((1, col, 1, col + span - 1))
            col += span
        xlsx_path = RESULTS_DIR / "results_compact.xlsx"
        lock = FileLock(str(xlsx_path) + ".lock")
        try:
            with lock.acquire(timeout=60):
                from openpyxl import load_workbook
                if not xlsx_path.exists():
                    _append_xlsx_with_headers(xlsx_path, [header1, header2], [], merges)
                # upsert into existing row keyed by (run_name, timestamp)
                wb = load_workbook(xlsx_path)
                ws = wb.active
                key_run = run_name
                key_ts = timestamp
                target_row = None
                for r in range(3, ws.max_row + 1):
                    if str(ws.cell(row=r, column=1).value or "") == key_run and str(ws.cell(row=r, column=2).value or "") == key_ts:
                        target_row = r
                        break
                if target_row is None:
                    target_row = ws.max_row + 1
                    # write config cols
                    for c_idx, val in enumerate(row[:len(header_config_cols)], start=1):
                        ws.cell(row=target_row, column=c_idx, value=val)
                # helper to compute group start
                base_col = len(header_config_cols) + 1
                group_span = 6
                def write_group(group_idx: int, fold_idx: Optional[int], mean_val: Optional[float], std_val: Optional[float], per_fold_vals: List[Optional[float]], digits: int = 4, cur_val: Optional[float] = None):
                    start = base_col + group_idx * group_span
                    if fold_idx is not None and 1 <= fold_idx <= 5:
                        val = per_fold_vals[fold_idx - 1]
                        if (val is None) and (cur_val is not None):
                            val = cur_val
                        if val is not None:
                            ws.cell(row=target_row, column=start + (fold_idx - 1), value=float(val) if digits > 0 else int(round(float(val))))
                    # recompute mean across fold cells 1..5 dynamically
                    vals = []
                    for c in range(0, 5):
                        cell = ws.cell(row=target_row, column=start + c).value
                        try:
                            if cell is not None and cell != "":
                                vals.append(float(cell))
                        except Exception:
                            pass
                    if vals:
                        mean_val_calc = sum(vals) / len(vals)
                        ws.cell(row=target_row, column=start + 5, value=float(mean_val_calc) if digits > 0 else int(round(float(mean_val_calc))))
                # per-fold updates
                write_group(0, fold_index, None, None, f1_vals, 4, cur_val=cur_f1)
                write_group(1, fold_index, None, None, iou_vals, 4, cur_val=cur_iou)
                write_group(2, fold_index, None, None, map_vals, 4, cur_val=cur_map)
                write_group(3, fold_index, None, None, fpr_vals, 4, cur_val=cur_fpr)
                write_group(4, fold_index, None, None, conf_vals, 3, cur_val=cur_conf)
                write_group(5, fold_index, None, None, epoch_vals, 0, cur_val=cur_epoch)
                wb.save(xlsx_path)
        except Timeout:
            raise RuntimeError(f"Timed out waiting for lock on {xlsx_path}")

    @staticmethod
    def append_extended(
        run_info: Dict[str, Any],
        aggregate: Dict[str, Any] | None,
        paths: Dict[str, Any] | None = None,
    ) -> None:
        """Append a row to results/results_extended.csv (single header row)."""
        ensure_dir(RESULTS_DIR)
        csv_path = RESULTS_DIR / "results_extended.csv"

        device_name, cuda_ver = _device_info()

        # No CSV header needed anymore for extended (XLSX only)

        # Build row
        run_name = str(run_info.get("run_name", run_info.get("run_id", "")))
        run_id = str(run_info.get("run_id", run_name))
        timestamp = run_info.get("timestamp") or _iso_utc_now()
        git_commit = run_info.get("git_commit") or _git_commit_short()

        model = str(run_info.get("model", ""))
        imgsz = run_info.get("imgsz")
        batch_size = run_info.get("batch_size")
        epochs_planned = run_info.get("epochs_planned")
        ema = run_info.get("ema")
        optimizer = run_info.get("optimizer")
        lr0 = run_info.get("lr0")
        lrf = run_info.get("lrf")
        weight_decay = run_info.get("weight_decay")
        momentum = run_info.get("momentum")
        scheduler = run_info.get("scheduler")
        n_augmentations = run_info.get("n_augmentations")
        op_iou_thresh = run_info.get("op_iou_thresh")
        fp_rate_cap = run_info.get("fp_rate_cap")
        max_det = run_info.get("max_det")
        edge_touch_k = run_info.get("edge_touch_k")
        conf_sweep_desc = run_info.get("conf_sweep_desc")
        seed = run_info.get("seed")

        # Aggregates from provided dict
        def g(key: str) -> Optional[float]:
            if not aggregate:
                return None
            v = aggregate.get(key)
            return float(v) if isinstance(v, (int, float)) else None

        row: List[str] = [
            run_name,
            run_id,
            timestamp,
            git_commit,
            device_name,
            cuda_ver,
            # core
            model,
            _fmt(imgsz, 0),
            _fmt(batch_size, 0),
            _fmt(epochs_planned, 0),
            str(bool(ema)) if ema is not None else "",
            str(optimizer) if optimizer is not None else "",
            format_float(lr0, ndigits=4),
            format_float(lrf, ndigits=4),
            format_float(weight_decay, ndigits=6) if weight_decay is not None else "",
            format_float(momentum, ndigits=4),
            str(scheduler) if scheduler is not None else "",
            _fmt(n_augmentations, 0),
            format_float(op_iou_thresh, ndigits=3),
            format_float(fp_rate_cap, ndigits=3),
            _fmt(max_det, 0),
            _fmt(edge_touch_k, 0),
            str(conf_sweep_desc) if conf_sweep_desc is not None else "",
            _fmt(seed, 0),
            # qualitative placeholders
            "","","",
            # aggregates
            format_float(g("f1_best_mean"), 4),
            format_float(g("f1_best_std"), 4),
            format_float(g("mean_iou_tp_best_mean"), 4) if aggregate and "mean_iou_tp_best_mean" in aggregate else format_float(g("mean_iou_tp_mean"), 4),
            format_float(g("mean_iou_tp_best_std"), 4) if aggregate and "mean_iou_tp_best_std" in aggregate else format_float(g("mean_iou_tp_std"), 4),
            format_float(g("map50_mean"), 4),
            format_float(g("map50_std"), 4),
            format_float(g("map50_95_mean"), 4),
            format_float(g("map50_95_std"), 4),
            format_float(g("best_conf_mean"), 3),
            format_float(g("best_conf_std"), 3),
            format_float(g("fp_rate_best_mean"), 4),
            _fmt(aggregate.get("best_epoch_mean") if aggregate else None, 0),
            str(paths.get("weights_best_path")) if paths and paths.get("weights_best_path") else "",
            str(paths.get("tb_dir")) if paths and paths.get("tb_dir") else "",
            str(paths.get("summary_json")) if paths and paths.get("summary_json") else "",
        ]

        # Only write XLSX (two-row header with group merges for readability)
        # Build group header row
        header1 = []
        header2 = []
        groups: List[tuple[str, List[str]]] = [
            ("Run identity", ["run_name","run_id","timestamp","git_commit","device_name","cuda_version"]),
            ("Core config", [
                "model","imgsz","batch_size","epochs_planned","ema",
                "optimizer","lr0","lrf","weight_decay","momentum","scheduler",
                "n_augmentations","op_iou_thresh","fp_rate_cap","max_det","edge_touch_k",
                "conf_sweep_desc","seed"
            ]),
            ("Qualitative", ["aug_profile","geom_profile","notes"]),
            ("Aggregates", [
                "f1_best_mean","f1_best_std",
                "mean_iou_tp_mean","mean_iou_tp_std",
                "map50_mean","map50_std",
                "map50_95_mean","map50_95_std",
                "best_conf_mean","best_conf_std",
                "fp_rate_best_mean",
                "best_epoch_mean",
                "weights_best_path","tb_dir","summary_json"
            ]),
        ]
        merges: List[tuple[int, int, int, int]] = []
        c = 1
        for gname, cols in groups:
            header1 += [gname] + [""] * (len(cols) - 1)
            header2 += cols
            merges.append((1, c, 1, c + len(cols) - 1))
            c += len(cols)
        xlsx_path = RESULTS_DIR / "results_extended.xlsx"
        lock = FileLock(str(xlsx_path) + ".lock")
        try:
            with lock.acquire(timeout=60):
                _append_xlsx_with_headers(xlsx_path, [header1, header2], row, merges)
        except Timeout:
            raise RuntimeError(f"Timed out waiting for lock on {xlsx_path}")


